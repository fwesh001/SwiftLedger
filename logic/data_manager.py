"""
Bulk data management for SwiftLedger.
Provides Excel template generation and member import.
"""

from __future__ import annotations

from datetime import date, datetime
from typing import Callable, Dict, List, Optional, Tuple

try:
    import pandas as pd
except Exception:  # pragma: no cover - handled at runtime
    pd = None

try:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
except Exception:  # pragma: no cover - handled at runtime
    Workbook = None
    Comment = None
    Alignment = None
    Font = None
    PatternFill = None
    DataValidation = None

from database.queries import add_member, get_member_by_staff_number


class BulkDataManager:
    TEMPLATE_HEADERS = [
        "Full Name",
        "Staff ID",
        "Phone",
        "Department",
        "Bank Name",
        "Account Number",
        "Date Joined (YYYY-MM-DD)",
        "Initial Savings",
        "Initial Loan",
    ]
    DEPARTMENT_OPTIONS = ["SLT", "Admin", "Teaching", "Non-Teaching"]

    def __init__(self, db_path: str = "swiftledger.db"):
        self.db_path = db_path

    def generate_import_template(self, filename: str) -> bool:
        if Workbook is None or Font is None or PatternFill is None or Alignment is None:
            return False

        assert Workbook is not None
        assert Font is not None
        assert PatternFill is not None
        assert Alignment is not None

        wb = Workbook()
        ws = wb.active
        ws.title = "Members"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(self.TEMPLATE_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        ws.freeze_panes = "A2"

        # Staff ID comment
        staff_id_cell = ws.cell(row=1, column=2)
        if Comment is not None:
            staff_id_cell.comment = Comment("Must be unique. Duplicate IDs will be skipped.", "SwiftLedger")

        # Department dropdown
        if DataValidation is not None:
            dept_list = ",".join(self.DEPARTMENT_OPTIONS)
            dv_dept = DataValidation(type="list", formula1=f'"{dept_list}"', allow_blank=True)
            ws.add_data_validation(dv_dept)
            dv_dept.add("D2:D2000")

            dv_account = DataValidation(type="textLength", operator="equal", formula1="10", allow_blank=True)
            ws.add_data_validation(dv_account)
            dv_account.add("F2:F2000")

        # Column formatting for phone and account number
        for row in range(2, 2001):
            ws.cell(row=row, column=3).number_format = "@"
            ws.cell(row=row, column=6).number_format = "@"

        # Dummy data in row 2
        dummy = [
            "John Doe",
            "SLT/001",
            "+2348012345678",
            "SLT",
            "UBA",
            "1234567890",
            "2024-01-01",
            5000,
            0,
        ]
        for col_idx, value in enumerate(dummy, start=1):
            ws.cell(row=2, column=col_idx, value=value)

        # Column widths
        widths = [24, 14, 16, 16, 14, 16, 20, 14, 12]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[self._column_letter(col_idx)].width = width

        wb.save(filename)
        return True

    def import_members_from_excel(
        self,
        filepath: str,
        progress_callback: Optional[Callable[[int, int], bool]] = None,
    ) -> Tuple[int, List[Dict]]:
        if pd is None:
            return 0, [{"row": 0, "name": "", "error": "pandas is not available."}]

        try:
            df = pd.read_excel(filepath, engine="openpyxl", dtype=str)
        except Exception as exc:
            return 0, [{"row": 0, "name": "", "error": f"Unable to read Excel file: {exc}"}]

        df = df.fillna("")

        ok, column_map, missing = self._map_columns(list(df.columns))
        if not ok:
            return 0, [{"row": 0, "name": "", "error": f"Invalid template. Missing columns: {', '.join(missing)}"}]

        df = df.rename(columns=column_map)

        total = len(df.index)
        success_count = 0
        error_log: List[Dict] = []

        for row_idx in range(len(df.index)):
            row = df.iloc[row_idx]
            row_num = row_idx + 2
            if progress_callback is not None:
                try:
                    keep_going = progress_callback(row_idx + 1, total)
                except Exception:
                    keep_going = True
                if keep_going is False:
                    error_log.append({"row": row_num, "name": "", "error": "Import cancelled by user."})
                    break

            try:
                full_name = self._clean_string(row.get("Full Name", ""))
                staff_id = self._clean_string(row.get("Staff ID", ""))
                phone = self._normalize_phone(row.get("Phone", ""))
                department = self._clean_string(row.get("Department", "")) or "SLT"
                bank_name = self._clean_string(row.get("Bank Name", "")) or "UBA"
                account_no = self._clean_string(row.get("Account Number", ""))
                date_joined = self._normalize_date(row.get("Date Joined (YYYY-MM-DD)", ""))
                initial_savings = self._to_float(row.get("Initial Savings", ""))
                initial_loan = self._to_float(row.get("Initial Loan", ""))

                if not full_name or not staff_id:
                    error_log.append({"row": row_num, "name": full_name, "error": "Full Name and Staff ID are required."})
                    continue

                exists_ok, existing = get_member_by_staff_number(self.db_path, staff_id)
                if exists_ok and existing:
                    error_log.append({"row": row_num, "name": full_name, "error": "Duplicate Staff ID."})
                    continue

                if account_no:
                    if not account_no.isdigit() or len(account_no) != 10:
                        error_log.append({"row": row_num, "name": full_name, "error": "Account Number must be 10 digits."})
                        continue

                if initial_savings is None or initial_loan is None:
                    error_log.append({"row": row_num, "name": full_name, "error": "Initial Savings/Loan must be numeric."})
                    continue

                member_data = {
                    "staff_number": staff_id,
                    "full_name": full_name,
                    "phone": phone,
                    "department": department,
                    "bank_name": bank_name,
                    "account_no": account_no,
                    "date_joined": date_joined,
                    "trans_date": date_joined,
                    "current_savings": initial_savings,
                    "total_loans": initial_loan,
                }

                ok_add, msg = add_member(self.db_path, member_data)
                if not ok_add:
                    error_log.append({"row": row_num, "name": full_name, "error": msg})
                    continue

                success_count += 1

            except Exception as exc:
                error_log.append({"row": row_num, "name": "", "error": f"Unexpected error: {exc}"})

        return success_count, error_log

    @staticmethod
    def export_error_log(error_log: List[Dict], filepath: str) -> bool:
        try:
            with open(filepath, "w", encoding="utf-8") as handle:
                handle.write("SwiftLedger Import Error Log\n")
                handle.write("================================\n\n")
                for err in error_log:
                    row = err.get("row", "?")
                    name = err.get("name", "")
                    error = err.get("error", "")
                    handle.write(f"Row {row}: {name} - {error}\n")
            return True
        except Exception:
            return False

    @staticmethod
    def _clean_string(value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value).strip()

    @staticmethod
    def _normalize_phone(value: object) -> str:
        phone = BulkDataManager._clean_string(value)
        if not phone:
            return "+234"

        if phone.startswith("+"):
            return phone

        if phone.startswith("234") and phone.isdigit():
            return f"+{phone}"

        digits = "".join(ch for ch in phone if ch.isdigit())
        if digits.startswith("0") and len(digits) >= 10:
            return "+234" + digits[1:]
        if len(digits) == 10:
            return "+234" + digits
        return phone

    @staticmethod
    def _to_float(value: object) -> Optional[float]:
        if value is None or value == "":
            return 0.0
        try:
            return float(str(value))
        except Exception:
            return None

    @staticmethod
    def _normalize_date(value: object) -> str:
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()

        text = BulkDataManager._clean_string(value)
        if not text:
            return date.today().isoformat()
        try:
            return datetime.strptime(text, "%Y-%m-%d").date().isoformat()
        except Exception:
            return date.today().isoformat()

    @classmethod
    def _map_columns(cls, columns: List[str]) -> Tuple[bool, Dict[str, str], List[str]]:
        normalized = {cls._normalize_header(col): col for col in columns}
        column_map: Dict[str, str] = {}
        missing: List[str] = []

        for header in cls.TEMPLATE_HEADERS:
            norm = cls._normalize_header(header)
            actual = normalized.get(norm)
            if not actual:
                missing.append(header)
                continue
            column_map[actual] = header

        return len(missing) == 0, column_map, missing

    @staticmethod
    def _normalize_header(text: str) -> str:
        return " ".join(str(text).strip().lower().split())

    @staticmethod
    def _column_letter(index: int) -> str:
        letters = ""
        while index:
            index, remainder = divmod(index - 1, 26)
            letters = chr(65 + remainder) + letters
        return letters
